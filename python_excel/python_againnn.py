# Fetch the data from the excel file using python module openpyxl
import openpyxl
# importing module
path = "data9.xlsx"
# Give the location of our file
data9_object = openpyxl.load_workbook(path)
# creating an object
data9_object = data9_object.active
# storing the data from active sheet into an object called data9_
# data9_object.title = "new"
data9_sheet = data9_object.title
get_data = data9_object.cell(row=1, column=1)

print(get_data.value)
print(data9_object.max_row)
print(data9_object.max_column)
print(data9_sheet)
data9_object.cell(row = 1, column = 1).value = "changed"
